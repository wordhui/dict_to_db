import re
import sys
import time
import json
import copy
import pickle
import logging
import sqlite3
import datetime
from threading import Lock
from typing import List, Union, Iterable, Callable, Generator, Tuple, Dict

from openpyxl import load_workbook, Workbook

CREATE_TABLE_SQL_TEMPLATE = f"create table{' '}[{{table_name}}] ({{column_info}});"
INSERT_SQL_TEMPLATE = f"insert into{' '}[{{table_name}}]({{columns}}) values({{values}});"
UPDATE_SQL_TEMPLATE = f"update{' '}[{{table_name}}] set {{update_column}} where {{where}};"
DELETE_SQL_TEMPLATE = f"delete from{' '}[{{table_name}}] where {{where}};"
SELECT_SQL_TEMPLATE = f"select {{select_column}} from{' '}[{{table_name}}] where {{where}};"
REPLACE_SQL_TEMPLATE = f"replace into{' '}[{{table_name}}]({{columns}}) values({{values}});"
INSERT_OR_UPDATE_SQL_TEMPLATE = f"replace into{' '}[{{table_name}}]({{columns}}) values({{values}});"
ADD_COLUMN_SQL_TEMPLATE = f"alter table{' '}[{{table_name}}] add {{column_info}};"
SELECT_TABLE_INDEX_NAMES = f"{'select'} name from MAIN.[sqlite_master] where type='index' and tbl_name=:table_name;"
PRAGMA_INDEX = "PRAGMA index_info({index_name});"
TABLE_TYPE_INFO = {str: 'text', int: 'integer', float: 'double', bool: 'boolean', datetime.date: 'date',
                   datetime.datetime: "timestamp", list: 'json_text', dict: 'json_text', tuple: 'tuple_text',
                   set: 'set_text'}
TABLE_COLUMN_SHORTHAND = {'pk': 'primary key', 'uq': 'unique'}
log = logging.getLogger("dict_to_db")
formatter = logging.Formatter('%(asctime)s %(levelname)-5s: %(message)s')
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(formatter)
log.addHandler(console_handler)
log.setLevel(logging.INFO)


def get_excel_title_by_index(index):
    """根据Excel 的title index 值获取Excel title名称"""
    name_list = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    title_list = []
    for i in range(1000):
        residue = index % 26
        index = index // 26
        title_list.insert(0, name_list[residue - 1])
        if index == 0:
            break
    return "".join(title_list)


def dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d


def adapt_obj(obj):
    return pickle.dumps(obj)


def convert_obj(obj_byte):
    return pickle.loads(obj_byte)


def convert_json_text(text):
    return json.loads(text)


def convert_tuple_text(text):
    return eval(text)


def convert_set_text(text):
    return eval(text)


# sqlite3.register_adapter(object, adapt_obj)
sqlite3.register_converter("obj", convert_obj)
sqlite3.register_converter("json_text", convert_json_text)
sqlite3.register_converter("tuple_text", convert_tuple_text)
sqlite3.register_converter("set_text", convert_set_text)


class DictToDb(object):
    def __init__(self, database: str = ":memory:", timeout: float = 5.0, detect_types: int = sqlite3.PARSE_DECLTYPES,
                 isolation_level: str = "DEFERRED", check_same_thread: bool = True,
                 cached_statements: int = 100, uri=False, row_factory: Callable = dict_factory,
                 insert_time: bool = True, update_time: bool = True, export: bool = True, auto_commit: bool = True,
                 auto_alter: bool = True):
        """
        :param database:数据库路径，也可以是 :memory: 表示这是一个内存数据库
        :param timeout:连接超时时间
        :param detect_types:默认为 0 (即关闭，不进行类型检测)，你可以将其设为任意的 PARSE_DECLTYPES 和 PARSE_COLNAMES 组合来启用类型检测
        :param isolation_level:事务隔离级别 可选值为 None(autocommit),"DEFERRED","IMMEDIATE","EXCLUSIVE"
        :param check_same_thread:是否只在一个线程中运行，默认为TRUE，若要多线程运行，请设置为FALSE
        :param cached_statements:缓存SQL语句的条数 默认100条
        :param uri:如果 uri 为真，则 database 被解释为 URI,它允许您指定选项。 例如，以只读模式打开数据库 sqlite3.connect('file:path/to/database?mode=ro', uri =True)
        :param row_factory:指定row_factory回调函数，默认的回调函数，会将查询出的结果行转换为dict,如果为了性能可以使用sqlite3.row替换默认的dict_factory,设置为none则返回为tuple类型
        :param insert_time 默认是否给表添加插入时间数据列，这里是全局设置，可以被方法内的insert_time参数局部覆盖
        :param update_time 默认是否给表添加更新时间数据列， 这里是全局设置，可以被方法内的update_time参数局部覆盖
        :param export 默认是否给表添加export数据列 ，这里是全局设置，可以被方法内的export参数局部覆盖
        :param auto_commit 是否自动执行commit语句，这里是全局设置，可以被方法内的commit参数局部覆盖
        :param auto_alter 是否自动执行alter 表结构，这里是全局设置，可以被方法内的auto_alter参数局部覆盖
        """
        self.db = sqlite3.connect(database, timeout=timeout, detect_types=detect_types, isolation_level=isolation_level,
                                  check_same_thread=check_same_thread, cached_statements=cached_statements,
                                  uri=uri)
        self._tables = {}
        self._insert_sql = {}
        self._insert_or_update_sql = {}
        self._update_sql = {}
        self._delete_sql = {}
        self._replace_sql = {}
        self._select_sql = {}
        self.lock = None
        self._insert_time = insert_time
        self._update_time = update_time
        self._export = export
        self._auto_commit = auto_commit
        self._auto_alter = auto_alter
        self._check_same_thread = check_same_thread
        self._re_pattern = {
            "pk": re.compile(r'primary\s+key$')
        }
        if not check_same_thread:
            self.lock = Lock()
        if row_factory:
            self.db.row_factory = row_factory
        self.cursor = self.db.cursor()
        self._load_db_tables()

    def insert(self, data: Union[dict, Iterable[dict], Generator[dict, None, None]], table_name: str = None,
               commit: bool = None, insert_time: bool = None, update_time: bool = None, export: bool = None,
               auto_alter: bool = None):
        """
        根据dict插入数据的函数，如果当前dict数据结构没有在数据库中建表，此函数则会自动建表
        :param data: 需要插入的dict 或者可迭代对象，且这个可迭代对象的子元素为dict
        :param table_name: 用户自定义表名，如果没有填写，则表名为t1,t2.....tn规则，依次递增
        :param commit: 是否插入一条语句后立即执行commit
        :param insert_time: 是否给数据加入一列插入时间列
        :param update_time: 是否给数据加入一列更新时间列
        :param export: 是否给数据加入一列导出数据列
        :param auto_alter: 是否自动alter表结构
        """
        if insert_time is None:
            insert_time = self._insert_time
        if update_time is None:
            update_time = self._update_time
        if export is None:
            export = self._export
        if commit is None:
            commit = self._auto_commit
        if auto_alter is None:
            auto_alter = self._auto_alter
        if isinstance(data, dict):
            insert_data = data
        elif isinstance(data, Generator):
            insert_data = next(data)
        elif isinstance(data, Iterable):
            for d in data:
                insert_data = d
                break
        else:
            raise Exception("不支持的类型 Unsupported type")
        if table_name is None:
            table_name = self._get_table_name_by_dict_keys(insert_data, insert_time, update_time, export)
        if table_name not in self._tables.keys():
            self._create_table_by_dict(insert_data, table_name, insert_time, update_time, export)
        insert_sql = self._get_insert_sql_by_dict(insert_data, table_name)
        try:
            self._execute_insert_sql(data, insert_data, insert_sql, table_name)
        except sqlite3.OperationalError as e:
            if auto_alter and (str(e).startswith("no such column") or 'no column named' in str(e)):
                self._alter_table_add_column_by_dict(insert_data, table_name=table_name)
                self._execute_insert_sql(data, insert_data, insert_sql, table_name)
            else:
                raise e
        self._commit(commit)

    def insert_or_update(self, data: Union[dict, Iterable[dict], Generator[dict, None, None]], table_name: str = None,
                         commit: bool = None, insert_time: bool = None, update_time: bool = None, export: bool = None,
                         auto_alter: bool = None, auto_update_time: bool = True, ignore_error=None):
        """
        不存在则插入，存在则更新的方法 【此方法性能略差于 insert 和insert_replace】
        根据dict插入数据的函数，如果当前dict数据结构没有在数据库中建表，此函数则会自动建表
        :param data: 需要插入的dict 或者可迭代对象，且这个可迭代对象的子元素为dict
        :param table_name: 用户自定义表名，如果没有填写，则表名为t1,t2.....tn规则，依次递增
        :param commit: 是否插入一条语句后立即执行commit
        :param insert_time: 是否给数据加入一列插入时间列
        :param update_time: 是否给数据加入一列更新时间列
        :param export: 是否给数据加入一列导出数据列
        :param auto_alter: 是否自动alter表结构
        :param auto_update_time: 是否在更新操作中自动更新update_time的值
        :param ignore_error: 是否在单次保存或更新中忽略某些异常以保证，数据大部分都插入到数据库中
        """
        if isinstance(data, dict):
            self._insert_or_update(data=data, table_name=table_name, commit=commit, insert_time=insert_time,
                                   update_time=update_time, export=export, auto_alter=auto_alter,
                                   auto_update_time=auto_update_time, ignore_error=ignore_error)
        elif isinstance(data, (Generator, Iterable)):
            for d in data:
                self._insert_or_update(data=d, table_name=table_name, commit=commit, insert_time=insert_time,
                                       update_time=update_time, export=export, auto_alter=auto_alter,
                                       auto_update_time=auto_update_time, ignore_error=ignore_error)
        else:
            raise Exception("不支持的类型 Unsupported type")

    def insert_or_replace(self, data: Union[dict, Iterable[dict], Generator[dict, None, None]], table_name: str = None,
                          commit: bool = None, insert_time: bool = None, update_time: bool = None, export: bool = None,
                          auto_alter: bool = True):
        """
        不存在则插入，存在则替换的函数，注：受sqlite replace方法的限制，如果某个字段没有填写被替换的值，这个值将会替换为Null
        根据dict插入或替换数据的函数，如果当前dict数据结构没有在数据库中建表，此函数则会自动建表
        :param data: 需要插入的dict 或者dict的list
        :param table_name: 用户自定义表名，如果没有填写，则表名为t1,t2.....tn规则，依次递增
        :param commit: 是否插入一条语句后立即执行commit
        :param insert_time: 是否给数据加入一列插入时间列
        :param update_time: 是否给数据加入一列更新时间列
        :param export: 是否给数据加入一列导出数据列
        :param auto_alter: 是否自动alter表结构
        """
        if insert_time is None:
            insert_time = self._insert_time
        if update_time is None:
            update_time = self._update_time
        if export is None:
            export = self._export
        if commit is None:
            commit = self._auto_commit
        if auto_alter is None:
            auto_alter = self._auto_alter
        if isinstance(data, dict):
            insert_data = data
        elif isinstance(data, Generator):
            insert_data = next(data)
        elif isinstance(data, Iterable):
            for d in data:
                insert_data = d
                break
        else:
            raise Exception("不支持的类型 Unsupported type")
        if table_name is None:
            table_name = self._get_table_name_by_dict_keys(insert_data, insert_time, update_time, export)
        if table_name not in self._tables.keys():
            create_table_sql = self._create_table_by_dict(insert_data, table_name, insert_time, update_time, export)
        replace_sql = self._get_replace_sql_by_dict(insert_data, table_name)
        try:
            self._execute_replace_sql(data, insert_data, replace_sql, table_name)
        except sqlite3.OperationalError as e:
            if auto_alter and (str(e).startswith("no such column") or 'no column named' in str(e)):
                self._alter_table_add_column_by_dict(insert_data, table_name=table_name)
                self._execute_replace_sql(data, insert_data, replace_sql, table_name)
            else:
                raise e
        self._commit(commit)

    def update(self, update: Union[dict, List[dict], Tuple[dict]], where: Union[dict, List[dict], Tuple[dict]],
               table_name: str, commit: bool = None, update_time: bool = None, auto_alter: bool = True):
        """
        简单的更新函数，根据传入的dict更新数据库的值，并可以选择自动填写update_time的值
        :param update: 需要更新的字段
        :param where: 更新条件
        :param table_name: 需要更新的表名
        :param commit: 是否自动commit，可以覆盖全局的commit设置
        :param update_time: 是否自动填写更新时间，可以覆盖全局的更新时间配置
        :param auto_alter: 是否自动alter表结构
        """
        if update_time is None:
            update_time = self._update_time
        if commit is None:
            commit = self._auto_commit
        if auto_alter is None:
            auto_alter = self._auto_alter
        if table_name not in self._tables.keys():
            raise Exception(f"no table by table_name:{table_name}")
        if isinstance(update, dict):
            self._update(update, where, table_name=table_name, update_time=update_time,
                         auto_alter=auto_alter)
        elif isinstance(update, (list, tuple)):
            if len(update) != len(where):
                raise Exception(f"update 和 where参数值不匹配")
            for count, data in enumerate(update):
                _where = where[count]
                self._update(data, _where, table_name=table_name, update_time=update_time,
                             auto_alter=auto_alter)
        self._commit(commit)

    def select(self, table_name: str, select: List[str] = None, where: dict = None, select_all: bool = True):
        """考虑到select 语句的方便程度，推荐使用 execute函数来执行查询语句,来实现更大的灵活性
        :param select:需要查询的列
        :param where: 查询条件
        :param table_name:表名
        :param select_all:是否查询所有，若为TRUE则返回所有数据，若为False则返回一条数据
        """
        select_sql = self._get_select_sql(table_name, select, where)
        if where:
            select_value = self._adapt_dict_value(where, table_name)
            result = self.execute(select_sql, select_value)
        else:
            result = self.execute(select_sql)
        if select_all:
            return result.fetchall()
        else:
            return result.fetchone()

    def delete(self, where: dict, table_name: str, commit: bool = None):
        """考虑到 delete语句的方便程度，推荐使用 execute函数来执行查询语句
        :param table_name:表名
        :param where:查询条件
        :param commit:是否立即提交
        """
        if commit is None:
            commit = self._auto_commit
        delete_sql = self._get_delete_sql(table_name, where)
        delete_value = self._adapt_dict_value(where, table_name)
        self.execute(delete_sql, delete_value)
        self._commit(commit)

    def excel_to_db(self, excel: str, table_names: Dict[str, str] = None, internal_table_name: bool = False,
                    transform_string: bool = True, title_to_column_name: bool = True,
                    title_row_index: Dict[str, int] = None, data_row_start_index: Dict[str, int] = None,
                    columns_desc: Dict[str, dict] = None,
                    columns_pretreatment_function: Dict[str, Dict[str, Callable]] = None,
                    appends_data: Dict[str, dict] = None, insert_time: bool = False, update_time: bool = False,
                    execute_func: str = 'insert', export: bool = False, ignore_error=None):
        """
        将结构比较单一Excel数据保存到数据库中，默认程序以Excel中每个有数据的sheet name为表名，
        每个sheet 内容行第一行为字段名，后续的[1:]行则会保存到数据库
        :param excel: Excel文件路径
        :param table_names: 覆盖默认的sheet name作为表名，指定为table_names里面的表名，例如：{"sheet1":"user"}将sheet1重命名为user表
        :param internal_table_name:系统自动设置表名,无视sheet name 和 table_names参数里面的值
        :param transform_string:是否将所有表格的值转化为字符串存储（数据库中字段类型即为text）， 若此参数为false（有的Excel格式混乱，可能出错）
                                ，则所有表值按照openpyxl读取为准，数据库内字段的类型以表格第二行每格数据的值的类型，来确定
        :param title_to_column_name: 是否根据Excel表格里面的表格标题名，做字段名，如果设置为false，则column名规则类似Excel表格从A-z 然后是AA-AZ 然后是BA-BZ...
        :param title_row_index: 每个sheet标题列所在的位置，默认为Excel中的第一列 如：{"sheet1":3} #sheet的title以第三行为准
        :param data_row_start_index: 每个sheet数据列起始位置，默认为Excel中的第二列开始 如：{"sheet1":20} # sheet1的数据从第20行开始保存
        :param columns_desc: 给指定的Excel列设置建表的描述信息，如设置为主键，unique等以及给某列设置数据库中存储类型 如: {'sheet1':{'A1':'@text#pk__not null__unique'}}
        :param columns_pretreatment_function: 给指定的Excel列 设置预处理函数，如 {'sheet1':{'A1':lambda x:float(x)}}
        :param appends_data: 插入Excel不包含的额外的数据列到数据库中，如给sheet1中的每行数据多插入一条 age数据：{'sheet1':{'age@text#pk':33}}
        :param insert_time: 是否添加插入时间列
        :param update_time：是否添加更新时间列
        :param export: 是否添加export数据列
        :param execute_func: 执行的方法，可选的有insert，replace，insert_or_update
        :param ignore_error: 是否在单次保存中忽略某些异常以保证，文件数据全部保存到Excel中
        """
        start_time = time.time()
        print(f"加载 {excel} 并保存到db中....")
        wb = load_workbook(excel, read_only=True)
        try:
            sheet_names = wb.sheetnames
            for sheet_count, sheet_name in enumerate(sheet_names):
                ws = wb[sheet_name]
                sheet_column_desc = columns_desc.get(sheet_name, {}) if columns_desc else {}
                sheet_append_data = appends_data.get(sheet_name, {}) if appends_data else {}
                sheet_title_index = title_row_index.get(sheet_name, 1) if title_row_index else 1
                sheet_data_row_start_index = data_row_start_index.get(sheet_name, 2) if data_row_start_index else 2
                sheet_columns_pretreatment_function = columns_pretreatment_function.get(sheet_name,
                                                                                        {}) if columns_pretreatment_function else {}
                column_names, create_table_dict, first_column_names = \
                    self._get_create_table_dict_by_excel_sheet(ws, title_to_column_name, sheet_title_index,
                                                               sheet_data_row_start_index, sheet_count,
                                                               transform_string,
                                                               sheet_column_desc, sheet_append_data)
                if not column_names:
                    continue
                if internal_table_name:
                    table_name = self._get_table_name_by_dict_keys(create_table_dict, insert_time, update_time, export)
                else:
                    table_name = sheet_name
                    if table_names and table_names.get(sheet_name):
                        table_name = table_names.get(sheet_name)
                for count, row in enumerate(ws.rows):
                    try:
                        if count >= sheet_data_row_start_index:
                            data = {}
                            blank_line = True  # 判断数据是不是全空
                            for cell_count, cell in enumerate(row):
                                cell_value = cell.value
                                first_column_name = first_column_names[cell_count]
                                columns_func = sheet_columns_pretreatment_function.get(
                                    first_column_name)
                                if cell_value is None:
                                    cell_value = ""
                                if transform_string and columns_func is None:
                                    cell_value = str(cell_value)
                                if cell_value:
                                    blank_line = False
                                    if columns_func:
                                        cell_value = columns_func(cell_value)
                                data[column_names[cell_count]] = cell_value
                            if not blank_line:
                                if appends_data:
                                    data.update(sheet_append_data)
                                if execute_func == "insert":
                                    self.insert(data, table_name, commit=False, insert_time=insert_time,
                                                update_time=update_time, export=export, auto_alter=False)
                                elif execute_func == "insert_or_update":
                                    self.insert_or_update(data, table_name, commit=False, insert_time=insert_time,
                                                          update_time=update_time, export=export, auto_alter=False)
                                elif execute_func == "replace":
                                    self.insert_or_replace(data, table_name, commit=False, insert_time=insert_time,
                                                           update_time=update_time, export=export, auto_alter=False)
                    except Exception as e:
                        if ignore_error and isinstance(e, ignore_error):
                            log.warning(e)
                        else:
                            raise e
                self._commit(commit=True)
            print(f"保存完成，共耗时：{round((time.time() - start_time), 2)} S")
        finally:
            wb.close()

    def excel_to_dict_list(self, excel: str, transform_string: bool = True, title_to_column_name: bool = True,
                           title_row_index: Dict[str, int] = None, data_row_start_index: Dict[str, int] = None,
                           columns_desc: Dict[str, dict] = None,
                           columns_pretreatment_function: Dict[str, Dict[str, Callable]] = None,
                           appends_data: Dict[str, dict] = None, ignore_error=None):
        """
        将结构比较单一Excel数据 转化为dict 格式返回，用生成器的方式
        每个sheet 内容行第一行为字段名，后续的[1:]行则会保存到数据库
        :param excel: Excel文件路径
        :param transform_string:是否将所有表格的值转化为字符串返回
        :param title_to_column_name: 是否根据Excel表格里面的表格标题名，做dict Key，如果设置为false，则Dict 的Key规则类似Excel表格从A-z 然后是AA-AZ 然后是BA-BZ...
        :param title_row_index: 每个sheet标题列所在的位置，默认为Excel中的第一列 如：{"sheet1":3} #sheet的title以第三行为准
        :param data_row_start_index: 每个sheet数据列起始位置，默认为Excel中的第二列开始 如：{"sheet1":20} # sheet1的数据从第20行开始保存
        :param columns_desc: 给指定的Excel列设置建表的描述信息，如设置为主键，unique等以及给某列设置数据库中存储类型 如: {'sheet1':{'A1':'@text#pk__not null__unique'}}
        :param columns_pretreatment_function: 给指定的Excel列 设置预处理函数，如 {'sheet1':{'A1':lambda x:float(x)}}
        :param appends_data: 插入Excel不包含的额外的数据列到数据库中，如给sheet1中的每行数据多插入一条 age数据：{'sheet1':{'age@text#pk':33}}
        :param ignore_error: 是否在单次保存中忽略某些异常以保证，文件数据全部保存到Excel中
        """
        start_time = time.time()
        print(f"加载 {excel} 并通过生成器方式返回 dict")
        wb = load_workbook(excel, read_only=True)
        try:
            sheet_names = wb.sheetnames
            for sheet_count, sheet_name in enumerate(sheet_names):
                ws = wb[sheet_name]
                sheet_column_desc = columns_desc.get(sheet_name, {}) if columns_desc else {}
                sheet_append_data = appends_data.get(sheet_name, {}) if appends_data else {}
                sheet_title_index = title_row_index.get(sheet_name, 1) if title_row_index else 1
                sheet_data_row_start_index = data_row_start_index.get(sheet_name, 2) if data_row_start_index else 2
                sheet_columns_pretreatment_function = columns_pretreatment_function.get(sheet_name,
                                                                                        {}) if columns_pretreatment_function else {}
                column_names, create_table_dict, first_column_names = \
                    self._get_create_table_dict_by_excel_sheet(ws, title_to_column_name, sheet_title_index,
                                                               sheet_data_row_start_index, sheet_count,
                                                               transform_string,
                                                               sheet_column_desc, sheet_append_data)
                if not column_names:
                    continue
                for count, row in enumerate(ws.rows):
                    try:
                        if count >= sheet_data_row_start_index:
                            data = {}
                            blank_line = True  # 判断数据是不是全空
                            for cell_count, cell in enumerate(row):
                                cell_value = cell.value
                                first_column_name = first_column_names[cell_count]
                                columns_func = sheet_columns_pretreatment_function.get(
                                    first_column_name)
                                if cell_value is None:
                                    cell_value = ""
                                if transform_string and columns_func is None:
                                    cell_value = str(cell_value)
                                if cell_value:
                                    blank_line = False
                                    if columns_func:
                                        cell_value = columns_func(cell_value)
                                data[column_names[cell_count]] = cell_value
                            if not blank_line:
                                if appends_data:
                                    data.update(sheet_append_data)
                                yield data
                    except Exception as e:
                        if ignore_error and isinstance(e, ignore_error):
                            log.warning(e)
                        else:
                            raise e
            print(f"加载完成，共耗时：{round((time.time() - start_time), 2)} S")
        finally:
            wb.close()

    def select_and_save_excel(self, sql: str, excel: str = None, transform_string: bool = True,
                              sql_value: Iterable = None, not_save_column: list = None,
                              auto_update_export: bool = False, update_export_by_column: List[str] = None,
                              update_export_table_name: str = None):
        """
        从数据库导出数据到Excel表格中
        :param sql: 查询的sql语句
        :param excel: Excel文件路径，如果Excel参数为None，则导出文件名格式为：f'dict_to_db_export_{datetime.now()}.xlsx'
        :param transform_string:是否将数据库中的值转化为字符串存储到Excel中
        :param sql_value:sql 位置参数的值
        :param not_save_column 查询出来字段中，不保存到Excel的字段
        :param auto_update_export 【不常见使用方法，可以不了解】是否自动更新导出后数据的export字段的值
        :param update_export_by_column 【不常见使用方法，可以不了解】根据哪些字段做where条件自动更新export的值
        :param update_export_table_name 【不常见使用方法，可以不了解】根据自动更新哪个表 export的值
        """
        not_save_column = set(not_save_column) if not_save_column else set()
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("sheet1")
        ws.freeze_panes = "A2"
        if sql_value:
            rows = self.execute(sql, sql_value).fetchall()
        else:
            rows = self.execute(sql, ).fetchall()
        update_data = []
        for count, row_data in enumerate(rows):
            if count == 0:
                ws.append([k for k in row_data.keys() if k not in not_save_column])
            if transform_string:
                ws.append([str(d) if d else '' for k, d in row_data.items() if k not in not_save_column])
            else:
                ws.append([d if d else '' for k, d in row_data.items() if k not in not_save_column])
            if auto_update_export:
                update_data.append({k: row_data[k] for k in update_export_by_column})
        if excel is None:
            excel = f'dict_to_db_export_{datetime.datetime.now().strftime("%Y-%m-%d %H时%M点%S分")}.xlsx'
        if rows:
            wb.save(excel)
            print(f"导出：{excel}")
            if auto_update_export:
                update = {'export': 1}
                updates = [update for i in range(len(update_data))]
                self.update(updates, where=update_data, table_name=update_export_table_name)
        else:
            print(f"当前查询无数据导出...")

    def get_table_sql_by_dict(self, data: dict, table_name: str = None, insert_time: bool = False,
                              update_time: bool = False, export: bool = False):
        """
        根据传入的dict参数,返回当前dict的建表语句
        :param data: 建表的dict
        :param table_name: 表名，可不填
        :param insert_time: 是否创建插入时间列
        :param update_time: 是否创建更新时间列
        :param export: 是否创建导出数据列
        :return: 返回建表语句
        """
        return self._create_table_by_dict(data, table_name, insert_time, update_time, export)

    def execute(self, sql: str, *args, **kwargs):
        """
        执行SQL语句，强烈推荐有占位符参数化SQL语句，如 execute("select * from t1 where name=? and age=?",['张三',18])
        :param sql:sql
        """
        if self._check_same_thread:
            return self.cursor.execute(sql, *args, **kwargs)
        else:
            try:
                self.lock.acquire(timeout=50)
                return self.cursor.execute(sql, *args, **kwargs)
            finally:
                self.lock.release()

    def executemany(self, sql: str, *args, **kwargs):
        """
        基于在序列 seq_of_parameters 中找到的所有形参序列或映射执行一条 SQL 命令 如
            execute("insert into t1(name,value) values(?,?);",[('张三',18),('李四',17),('王五',16)])
        :param sql:sql
        """
        if self._check_same_thread:
            return self.cursor.executemany(sql, *args, **kwargs)
        else:
            try:
                self.lock.acquire(timeout=50)
                return self.cursor.executemany(sql, *args, **kwargs)
            finally:
                self.lock.release()

    def create_function(self, name: str, num_params: int, func: Callable, deterministic: bool = False):
        """
        创建一个可以在 SQL 语句中使用的用户自定义函数
        :param name:函数名
        :param num_params:该函数所接受的形参个数（如果 num_params 为 -1，则该函数可接受任意数量的参数）
        :param func:func 是一个 Python 可调用对象，它将作为 SQL 函数被调用
        :param deterministic:如果 deterministic 为真值，则所创建的函数将被标记为 deterministic，这允许 SQLite 执行额外的优化。 此旗标在 SQLite 3.8.3 或更高版本中受到支持，如果在旧版本中使用将引发 NotSupportedError
        """
        if self._check_same_thread:
            self.db.create_function(name, num_params, func, deterministic=deterministic)
        else:
            try:
                self.lock.acquire(timeout=50)
                self.db.create_function(name, num_params, func, deterministic=deterministic)
            finally:
                self.lock.release()
        # self.cursor = self.db.cursor() #这里不用执行这行语句也能生效

    def commit(self):
        """
        给外层用户使用的commit函数
        """
        if self._check_same_thread:
            self.db.commit()
        else:
            try:
                self.lock.acquire(timeout=50)
                self.db.commit()
            finally:
                self.lock.release()

    def executescript(self, sql: str):
        if self._check_same_thread:
            return self.cursor.executescript(sql)
        else:
            try:
                self.lock.acquire(timeout=50)
                return self.cursor.executescript(sql)
            finally:
                self.lock.release()

    def close(self):
        """
        执行commit后关闭数据库连接
        """
        if self._check_same_thread:
            self.db.commit()
            self.cursor.close()
            self.db.close()
        else:
            try:
                self.lock.acquire(timeout=50)
                self.db.commit()
                self.cursor.close()
                self.db.close()
            finally:
                self.lock.release()

    def _commit(self, commit: bool):
        """
        给函数内部使用的commit函数
        """
        if commit:
            self.commit()

    def _load_db_tables(self):
        """
        从数据库加载表结构
        """
        table_name_infos = self.db.execute("Select name From MAIN.[sqlite_master] where type='table';").fetchall()
        table_names = [t['name'] for t in table_name_infos]
        for table_name in table_names:
            # 下面这条语句不支持？占位符和命名占位符   不知道原因
            table_info_dict = {}
            table_info = self.db.execute(f"PRAGMA table_info([{table_name}]);").fetchall()
            for column_info in table_info:
                table_info_dict[column_info['name']] = column_info
            self._tables[table_name] = table_info_dict

    def _alter_table_add_column_by_dict(self, data: dict, table_name: str):
        self._load_db_tables()
        alter_table_sql = ""
        for key, value in data.items():
            if "@" in key:
                key = key.split("@")[0]
            elif "#" in key:
                key = key.split("#")[0]
            if key not in self._tables[table_name].keys():
                column_info_dict = self._get_column_info_by_key_value(key, value)
                column_info, pk_column = column_info_dict['column_info'], column_info_dict['pk_column']
                add_column_sql = ADD_COLUMN_SQL_TEMPLATE.format(table_name=table_name, column_info=column_info)
                alter_table_sql += f"{add_column_sql}\n"
                if pk_column:
                    raise Exception("不支持带主键的自动alter")
        self.executescript(alter_table_sql)
        self.commit()

    def _get_table_name_by_dict_keys(self, data: dict, insert_time: bool, update_time: bool, export: bool):
        """根据dict key值获取表名"""
        dict_column = set()
        for column in data.keys():
            if not isinstance(column, str):
                raise Exception('cn:需要保存到数据库的dict的key必须可以转化为字符串\nen:The key of the dict that '
                                'needs to be saved to the database must be convertible into a string')
            if '@' in column:
                dict_column.add(column.split('@')[0])
            elif '#' in column:
                dict_column.add(column.split('#')[0])
            else:
                dict_column.add(column)
        if insert_time:
            dict_column.add('insert_time')
        if update_time:
            dict_column.add("update_time")
        if export:
            dict_column.add('export')
        for table_name, table_column_infos in self._tables.items():
            table_columns = set(table_column_infos.keys())
            if set(table_columns) == dict_column:
                return table_name
        table_names = self._tables.keys()
        table_name_count = 1
        for table_name in table_names:
            result = "".join(re.findall(r'^t(\d+)$', table_name)).strip()
            if result and int(result) >= table_name_count:
                table_name_count = int(result) + 1
        return f"t{table_name_count}"

    def _create_table_by_dict(self, data: dict, table_name: str, insert_time: bool, update_time: bool,
                              export: bool) -> str:
        """
        根据dict结构拼接创建数据表的SQL，并执行SQL，最后返回建表的SQL
        :param data: 传入的dict
        :param table_name: 表名
        :param insert_time: 是否创建插入时间数据列
        :param update_time: 是否创建更新时间数据列
        :param export: 是否创建export睡觉列
        :return: 建表的SQL
        """
        column_info_list = []
        pk_column_list = []
        for key, value in data.items():
            column_info_dict = self._get_column_info_by_key_value(key, value)
            column_info, pk_column = column_info_dict['column_info'], column_info_dict['pk_column']
            column_info_list.append(column_info)
            if pk_column:
                pk_column_list.append(pk_column)
        if insert_time:
            column_info_list.append("insert_time timestamp default (datetime('now','localtime'))")
        if update_time:
            column_info_list.append('update_time timestamp')
        if export:
            column_info_list.append('export boolean default false')
        if pk_column_list:
            column_info_list.append(f"primary key ({','.join(pk_column_list)})")
        column_info = ", ".join(column_info_list)
        create_table_sql = CREATE_TABLE_SQL_TEMPLATE.format(table_name=table_name, column_info=column_info)
        self.cursor.execute(create_table_sql)
        self.db.commit()
        self._load_db_tables()
        return create_table_sql

    def _get_column_info_by_key_value(self, key, value) -> dict:
        """
        根据dict key和value 生成column info信息，及返回当前字段是不是主键，若是主键，则返回主键名称
        :param key:字典的key
        :param value:字典的value
        :return:dict(column_info：column_info信息,pk_column：是否是主键字段，None则不是主键column,str则表示为主键column，
                且主键名为该str,column_type 字段类型)
        """
        column_name = key
        column_desc = ""
        pk_column = None
        if '@' in key:
            column_name = column_name.split("@")[0]
            column_type = "".join(re.findall(r'@(\w+)[#]*', key))
        elif isinstance(value, (str, int, float, bool, datetime.date, datetime.datetime)):
            column_type = TABLE_TYPE_INFO[type(value)]
        elif isinstance(value, (list, dict, set, tuple)):
            if value == eval(str(value)):
                column_type = TABLE_TYPE_INFO[type(value)]
            else:
                column_type = "obj"
        elif isinstance(value, object):
            column_type = 'obj'
        else:
            raise TypeError("cn：不支持的存储类型\nen：Storage type not supported")
        if '#' in key:
            if '@' not in key:
                column_name = key.split("#")[0]
            column_desc_info = key.split("#")[1].split("__")
            column_desc_list = []
            for c_desc in column_desc_info:
                if ';' in c_desc:
                    raise Exception("column描述信息里面不应该包含字符';' ")
                if c_desc == "pk" or self._re_pattern['pk'].match(c_desc):
                    pk_column = f"[{column_name}]"
                elif c_desc in TABLE_COLUMN_SHORTHAND:
                    column_desc_list.append(TABLE_COLUMN_SHORTHAND[c_desc])
                else:
                    column_desc_list.append(c_desc)
            column_desc = " ".join(column_desc_list)
        column_name = f"[{column_name}]"
        column_info = " ".join([column_name, column_type, column_desc])
        return {"column_info": column_info, "pk_column": pk_column, "column_type": column_type}

    def _get_insert_sql_by_dict(self, data: dict, table_name: str) -> str:
        """
        根据传入的字典和表名拼接 插入的SQL语句
        """
        insert_sql_key = f"{'-'.join(data.keys())}_{table_name}"
        if insert_sql_key in self._insert_sql:
            return self._insert_sql[insert_sql_key]
        insert_column_names = []
        for column in data.keys():
            if '@' in column:
                insert_column_names.append(f"[{column.split('@')[0]}]")
            elif '#' in column:
                insert_column_names.append(f"[{column.split('#')[0]}]")
            else:
                insert_column_names.append(f"[{column}]")
        columns = ", ".join(insert_column_names)
        values = ",".join(['?' for i in range(len(insert_column_names))])
        insert_sql = INSERT_SQL_TEMPLATE.format(table_name=table_name, columns=columns, values=values)
        self._insert_sql[insert_sql_key] = insert_sql
        return insert_sql

    def _get_insert_or_update_sql_by_dict(self, data: dict, table_name: str) -> str:
        """
        根据传入的字典，和表名，拼接不存在则插入，存在则更新的SQL
        """
        insert_or_update_sql_key = f"{'-'.join(data.keys())}_{table_name}"
        if insert_or_update_sql_key in self._insert_or_update_sql:
            return self._insert_or_update_sql[insert_or_update_sql_key]
        insert_column_names = []
        for column in data.keys():
            if '@' in column:
                insert_column_names.append(f"[{column.split('@')[0]}]")
            elif '#' in column:
                insert_column_names.append(f"[{column.split('#')[0]}]")
            else:
                insert_column_names.append(f"[{column}]")
        columns = ", ".join(insert_column_names)
        values = ",".join(['?' for i in range(len(insert_column_names))])
        insert_or_update_sql = INSERT_SQL_TEMPLATE.format(table_name=table_name, columns=columns, values=values)
        self._insert_or_update_sql[insert_or_update_sql_key] = insert_or_update_sql
        return insert_or_update_sql

    @staticmethod
    def _get_update_data_by_where_column(insert_data, where_column):
        """根据where column 自动从inset data中获取update data 和 where data"""
        insert_column_data = {}
        for column, value in insert_data.items():
            if '@' in column:
                insert_column_data[column.split('@')[0]] = value
            elif '#' in column:
                insert_column_data[column.split('#')[0]] = value
            else:
                insert_column_data[column] = value
        update_data = {key: value for key, value in insert_column_data.items() if key not in where_column}
        where_data = {key: value for key, value in insert_column_data.items() if key in where_column}
        return update_data, where_data

    def _get_replace_sql_by_dict(self, data: dict, table_name: str) -> str:
        """
        根据传入的字典和表名拼接 插入的SQL语句
        """
        replace_sql_key = f"{'-'.join(data.keys())}_{table_name}"
        if replace_sql_key in self._replace_sql:
            return self._replace_sql[replace_sql_key]
        replace_column_names = []
        for column in data.keys():
            if '@' in column:
                replace_column_names.append(f"[{column.split('@')[0]}]")
            elif '#' in column:
                replace_column_names.append(f"[{column.split('#')[0]}]")
            else:
                replace_column_names.append(f"[{column}]")
        columns = ", ".join(replace_column_names)
        values = ",".join(['?' for i in range(len(replace_column_names))])
        replace_sql = REPLACE_SQL_TEMPLATE.format(table_name=table_name, columns=columns, values=values)
        self._replace_sql[replace_sql_key] = replace_sql
        return replace_sql

    def _execute_insert_sql(self, data, insert_data, insert_sql, table_name):
        """处理insert 函数SQL语句执行部分"""
        if isinstance(data, Generator):  # 处理生成器的情况，生成器需要单独保存第一次的状态
            self.execute(insert_sql, self._adapt_dict_value(insert_data, table_name))
            self.executemany(insert_sql, self._adapt_dict_values(data, table_name))
        elif isinstance(data, dict):
            self.execute(insert_sql, self._adapt_dict_value(data, table_name))
        elif isinstance(data, Iterable):
            self.executemany(insert_sql, self._adapt_dict_values(data, table_name))

    def _execute_replace_sql(self, data, insert_data, replace_sql, table_name):
        """处理insert_or_replace 函数SQL语句执行部分"""
        if isinstance(data, Generator):  # 处理生成器的情况，生成器需要单独保存第一次的状态
            self.execute(replace_sql, self._adapt_dict_value(insert_data, table_name))
            self.executemany(replace_sql, self._adapt_dict_values(data, table_name))
        elif isinstance(data, dict):
            self.execute(replace_sql, self._adapt_dict_value(data, table_name))
        elif isinstance(data, Iterable):
            self.executemany(replace_sql, self._adapt_dict_values(data, table_name))

    def _insert_or_update(self, data: Union[dict, Iterable[dict], Generator[dict, None, None]], table_name: str = None,
                          commit: bool = None, insert_time: bool = None, update_time: bool = None, export: bool = None,
                          auto_alter: bool = None, auto_update_time: bool = True, ignore_error=None):
        """insert or update 函数的调用逻辑"""
        try:
            self.insert(data, table_name=table_name, commit=commit, insert_time=insert_time,
                        update_time=update_time, export=export, auto_alter=auto_alter)
        except sqlite3.IntegrityError as e:
            if 'UNIQUE constraint failed' in str(e):
                where_column = str(e).replace("UNIQUE constraint failed: ", "").replace(f"{table_name}.",
                                                                                        '').split(", ")
                update_data, where_data = self._get_update_data_by_where_column(data, where_column)
                self.update(update_data, where_data, table_name=table_name, commit=commit, update_time=update_time,
                            auto_alter=auto_alter)
            else:
                raise e
        except Exception as e:
            if ignore_error and isinstance(e, ignore_error):
                log.warning(e)
            else:
                raise e

    def _update(self, update: Union[dict, List[dict], Tuple[dict]], where: Union[dict, List[dict], Tuple[dict]],
                table_name: str, update_time: bool = None, auto_alter: bool = True):
        """update 函数的执行逻辑"""
        update_sql = self._get_update_sql(update_data=update, where=where, table_name=table_name,
                                          update_time=update_time)
        update_values = self._get_update_column_and_where_values(update, where, update_time, table_name)
        try:
            self.execute(update_sql, update_values)
        except sqlite3.OperationalError as e:
            if auto_alter and (str(e).startswith("no such column") or 'no column named' in str(e)):
                self._alter_table_add_column_by_dict(update, table_name=table_name)
                self.execute(update_sql, update_values)
            else:
                raise e

    def _get_update_sql(self, update_data: dict, where: dict, table_name: str, update_time: bool):
        """根据传入的参数，拼接更新的SQL语句"""
        update_sql_key = f'{"-".join(update_data.keys())}@{"-".join(where.keys())}_{table_name}'
        if update_sql_key in self._update_sql.keys():
            return self._update_sql[update_sql_key]
        update_column_names = []
        where_column_names = []
        for column in update_data.keys():
            if '@' in column:
                update_column_names.append(f"[{column.split('@')[0]}]=?")
            elif '#' in column:
                update_column_names.append(f"[{column.split('#')[0]}]=?")
            else:
                update_column_names.append(f"[{column}]=?")
        if update_time and 'update_time' not in update_data.keys():
            update_column_names.append("update_time=?")
        for column in where.keys():
            if '@' in column:
                where_column_names.append(f"[{column.split('@')[0]}]=?")
            elif '#' in column:
                where_column_names.append(f"[{column.split('#')[0]}]=?")
            else:
                where_column_names.append(f"[{column}]=?")
        update_sql = UPDATE_SQL_TEMPLATE.format(table_name=table_name,
                                                update_column=",".join(update_column_names),
                                                where=" and ".join(where_column_names))
        self._update_sql[update_sql_key] = update_sql
        return update_sql

    @staticmethod
    def _get_select_sql(table_name: str, select: list, where: dict):
        if isinstance(select, list):
            select = f'[{"],[".join(select)}]'
        elif select is None:
            select = "*"
        if isinstance(where, dict):
            where = f"{'=?,'.join(where.keys())}=?"
        elif where is None:
            where = "1=1"
        select_sql = SELECT_SQL_TEMPLATE.format(select_column=select, table_name=table_name, where=where)
        return select_sql

    @staticmethod
    def _get_delete_sql(table_name: str, where: dict):
        return DELETE_SQL_TEMPLATE.format(table_name=table_name, where=f"{'=?,'.join(where.keys())}=?")

    def _get_update_column_and_where_values(self, update_data: dict, where: dict, update_time: bool, table_name: str):
        if update_time and 'update_time' not in update_data.keys():
            update_data['update_time'] = datetime.datetime.now()
        update_values = self._adapt_dict_value(update_data, table_name)
        where_values = self._adapt_dict_value(where, table_name)
        result_values = update_values + where_values
        return result_values

    @staticmethod
    def _get_excel_insert_data_by_generator(ws, data_row_start_index, sheet_count, transform_string,
                                            column_names):
        for count, row in enumerate(ws.rows):
            if (data_row_start_index is None and count >= 1) or (
                    data_row_start_index and count >= data_row_start_index[sheet_count]):
                data = {}
                blank_line = True  # 判断数据是不是全空
                for cell_count, cell in enumerate(row):
                    cell_value = cell.value
                    if cell_value is None:
                        cell_value = ""
                    if transform_string:
                        cell_value = str(cell_value)
                    if cell_value:
                        blank_line = False
                    data[column_names[cell_count]] = cell_value
                if not blank_line:
                    yield data

    @staticmethod
    def _get_create_table_dict_by_excel_sheet(ws, title_to_column_name, title_row_index, data_row_start_index,
                                              sheet_count, transform_string, sheet_column_desc, sheet_append_data):
        column_names = []
        column_values = []
        for count, row in enumerate(ws.rows):
            if title_row_index and count == title_row_index - 1:
                if title_to_column_name:
                    for cell_count, cell in enumerate(row):
                        col_name = cell.value
                        if col_name is not None:
                            column_names.append(str(col_name))
                        else:
                            column_names.append(get_excel_title_by_index(cell_count + 1))
                else:
                    column_names = [get_excel_title_by_index(_ + 1) for _ in range(len(row))]
            elif count == data_row_start_index - 1:
                for cell_count, cell in enumerate(row):
                    cell_value = cell.value
                    if cell_value is None:
                        cell_value = ""
                    if transform_string:
                        cell_value = str(cell_value)
                    column_values.append(cell_value)
            if column_names and len(column_names) == len(column_values):
                break  # 这里说明准确获取到column name 和column values了
        first_column_names = copy.deepcopy(column_names)
        if sheet_column_desc and column_names:
            for k, v in sheet_column_desc.items():
                for c, column in enumerate(column_names):
                    if column == k:
                        column_names[c] = column + v
        create_table_dict = {column_names[i]: column_values[i] for i in range(len(column_names))}
        if sheet_append_data and column_names:
            create_table_dict.update(sheet_append_data)
        return column_names, create_table_dict, first_column_names

    def _adapt_dict_value(self, data: dict, table_name: str):
        """
        将dict的值转为SQLite存储的的值
        """
        result_data = []
        for column, value in data.items():
            if isinstance(value, (str, int, float, bool, datetime.date, datetime.datetime)):
                result_data.append(value)
            elif value is None:
                result_data.append(None)
            else:
                try:
                    column_type = self._tables[table_name][column]['type']
                except KeyError:
                    column_name = column
                    if '@' in column:
                        column_name = column.split('@')[0]
                    elif '#' in column:
                        column_name = column.split('#')[0]
                    try:
                        column_type = self._tables[table_name][column_name]
                    except KeyError:  # 这里表明有表里不存在的字段
                        column_type = self._get_column_info_by_key_value(column, value)['column_type']
                if column_type == "json_text":  # 这里会将字典里面的tuple值转为list
                    result_data.append(json.dumps(value, ensure_ascii=False))
                elif column_type in ['tuple_text', 'set_text']:
                    result_data.append(str(value))
                elif column_type == "obj":
                    result_data.append(pickle.dumps(value))
        return result_data

    def _adapt_dict_values(self, data_list: Iterable[dict], table_name: str):
        """
        采用生成器方式，将data_list里面的每一项转为与SQLite交流的值
        """
        for data in data_list:
            yield self._adapt_dict_value(data, table_name)


__all__ = ['DictToDb']

if __name__ == '__main__':
    pass
